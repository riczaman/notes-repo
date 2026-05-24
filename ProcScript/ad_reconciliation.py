"""
================================================================================
AD Group Membership Reconciliation Tool
Enterprise-Grade | Banking Environment
================================================================================
Purpose : Compare Active Directory group members against a monthly Excel roster
Output  : Formatted Excel reconciliation workbook with full audit trail
Author  : Infrastructure Automation
Version : 2.0
================================================================================
"""

# ──────────────────────────────────────────────────────────────────────────────
# SECTION 1 ─ CONFIGURATION  (edit these values before running)
# ──────────────────────────────────────────────────────────────────────────────

CONFIG = {

    # ── Active Directory ──────────────────────────────────────────────────────
    "AD_GROUP_NAME": "GRP-BANKING-USERS",          # Target AD group (exact name)
    "AD_QUERY_TIMEOUT_SECONDS": 300,                # PowerShell timeout (5 min)
    "AD_DOMAIN_CONTROLLER": "",                     # Optional: "" = auto-detect

    # ── AD → DataFrame column names produced by the PowerShell export ─────────
    # These are the column headers that come out of Get-ADGroupMember/Get-ADUser.
    # Change only if your AD schema uses different property names.
    "AD_COLUMN_SAMACCOUNTNAME": "SamAccountName",  # AD login name
    "AD_COLUMN_MAIL":           "Mail",             # AD email field
    "AD_COLUMN_DISPLAYNAME":    "DisplayName",      # AD display name
    "AD_COLUMN_EMPLOYEEID":     "EmployeeID",       # AD employee ID (may be empty)

    # ── Excel input file ──────────────────────────────────────────────────────
    "INPUT_FOLDER":   "input",                      # Sub-folder containing the monthly file
    "EXCEL_SHEET_NAME": 0,                          # Sheet name or 0-based index (0 = first sheet)

    # ── Excel input → column name mapping ────────────────────────────────────
    # Set each value to the EXACT column header that appears in your monthly
    # Excel file.  Use None if the column does not exist in your file.
    "EXCEL_COLUMN_USERID":      "UserID",           # Primary match key  ← KEY VARIABLE
    "EXCEL_COLUMN_EMPLOYEEID":  "EmployeeID",       # Secondary match key (or None)
    "EXCEL_COLUMN_EMAIL":       "Email",            # Email column        (or None)
    "EXCEL_COLUMN_DISPLAYNAME": "DisplayName",      # Display name        (or None)

    # ── Comparison / matching ─────────────────────────────────────────────────
    # Which AD column maps to which Excel column for the PRIMARY comparison.
    # The value on the left must be one of the AD_COLUMN_* keys above.
    # The value on the right must be one of the EXCEL_COLUMN_* keys above.
    #
    # Example A – match on sAMAccountName vs UserID:
    #   "MATCH_AD_FIELD":    "AD_COLUMN_SAMACCOUNTNAME"
    #   "MATCH_EXCEL_FIELD": "EXCEL_COLUMN_USERID"
    #
    # Example B – match on EmployeeID vs EmployeeID:
    #   "MATCH_AD_FIELD":    "AD_COLUMN_EMPLOYEEID"
    #   "MATCH_EXCEL_FIELD": "EXCEL_COLUMN_EMPLOYEEID"
    #
    "MATCH_AD_FIELD":    "AD_COLUMN_SAMACCOUNTNAME",  # AD field used for matching
    "MATCH_EXCEL_FIELD": "EXCEL_COLUMN_USERID",       # Excel field used for matching

    # Fallback match fields tried in order when primary match finds no result.
    # Each entry is a tuple: (AD_COLUMN_* key, EXCEL_COLUMN_* key)
    # Set to [] to disable fallback matching.
    "FALLBACK_MATCH_PAIRS": [
        ("AD_COLUMN_EMPLOYEEID",     "EXCEL_COLUMN_EMPLOYEEID"),
        ("AD_COLUMN_MAIL",           "EXCEL_COLUMN_EMAIL"),
    ],

    # ── Output ────────────────────────────────────────────────────────────────
    "OUTPUT_FOLDER":        "reports",              # Where reconciliation reports are saved
    "ARCHIVE_FOLDER":       "archive",              # Old reports moved here automatically
    "LOG_FOLDER":           "logs",                 # Log files location
    "EXPORT_CSV_COPIES":    True,                   # Also write CSV per-tab
    "ARCHIVE_OLD_REPORTS":  True,                   # Move previous reports to archive/
    "DRY_RUN":              False,                  # True = analyse only, no files written

    # ── Behaviour ─────────────────────────────────────────────────────────────
    "CASE_SENSITIVE_MATCH": False,                  # False = case-insensitive comparison
    "LOG_LEVEL":            "INFO",                 # DEBUG | INFO | WARNING | ERROR
}

# ──────────────────────────────────────────────────────────────────────────────
# SECTION 2 ─ IMPORTS
# ──────────────────────────────────────────────────────────────────────────────

import glob
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ──────────────────────────────────────────────────────────────────────────────
# SECTION 3 ─ LOGGING SETUP
# ──────────────────────────────────────────────────────────────────────────────

def setup_logging(log_folder: str, level: str) -> logging.Logger:
    """Initialise structured logging to both console and a dated log file."""
    Path(log_folder).mkdir(parents=True, exist_ok=True)
    log_file = Path(log_folder) / f"ad_recon_{datetime.now():%Y%m%d_%H%M%S}.log"

    numeric_level = getattr(logging, level.upper(), logging.INFO)
    logger = logging.getLogger("ADRecon")
    logger.setLevel(numeric_level)

    fmt = logging.Formatter(
        fmt="%(asctime)s | %(levelname)-8s | %(funcName)-30s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(fmt)

    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.info("Logging initialised → %s", log_file)
    return logger


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 4 ─ ACTIVE DIRECTORY QUERY
# ──────────────────────────────────────────────────────────────────────────────

POWERSHELL_SCRIPT = r"""
param (
    [string]$GroupName,
    [string]$DomainController
)

$ErrorActionPreference = "Stop"

try {
    $params = @{ Identity = $GroupName; Recursive = $true }
    if ($DomainController -ne "") { $params["Server"] = $DomainController }

    $members = Get-ADGroupMember @params | Where-Object { $_.objectClass -eq "user" }

    $results = foreach ($member in $members) {
        $userParams = @{
            Identity   = $member.distinguishedName
            Properties = @("mail", "displayName", "employeeID", "enabled")
        }
        if ($DomainController -ne "") { $userParams["Server"] = $DomainController }

        try {
            $user = Get-ADUser @userParams
            [PSCustomObject]@{
                SamAccountName = $user.SamAccountName
                Mail           = $user.mail
                DisplayName    = $user.DisplayName
                EmployeeID     = $user.employeeID
                Enabled        = $user.Enabled
            }
        } catch {
            Write-Warning "Could not retrieve details for $($member.SamAccountName): $_"
        }
    }

    $results | ConvertTo-Json -Depth 3 -Compress
} catch {
    Write-Error "AD query failed: $_"
    exit 1
}
"""


def query_active_directory(cfg: dict, logger: logging.Logger) -> pd.DataFrame:
    """
    Execute a PowerShell recursive AD group query and return a DataFrame.
    Uses a temporary .ps1 file to avoid shell-escaping issues.
    """
    logger.info("Starting AD query for group: %s", cfg["AD_GROUP_NAME"])
    t0 = time.perf_counter()

    tmp_ps1 = Path("_ad_query_tmp.ps1")
    try:
        tmp_ps1.write_text(POWERSHELL_SCRIPT, encoding="utf-8")

        cmd = [
            "powershell.exe",
            "-NonInteractive",
            "-NoProfile",
            "-ExecutionPolicy", "Bypass",
            "-File", str(tmp_ps1),
            "-GroupName", cfg["AD_GROUP_NAME"],
            "-DomainController", cfg.get("AD_DOMAIN_CONTROLLER", ""),
        ]

        logger.info("Executing PowerShell (timeout=%ss)…", cfg["AD_QUERY_TIMEOUT_SECONDS"])
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=cfg["AD_QUERY_TIMEOUT_SECONDS"],
        )

        if result.returncode != 0:
            raise RuntimeError(
                f"PowerShell exited with code {result.returncode}.\n"
                f"STDERR:\n{result.stderr.strip()}"
            )

        stderr_stripped = result.stderr.strip()
        if stderr_stripped:
            for line in stderr_stripped.splitlines():
                logger.warning("PS STDERR: %s", line)

        stdout = result.stdout.strip()
        if not stdout:
            raise ValueError("PowerShell returned no output. Group may be empty or name is wrong.")

        raw = json.loads(stdout)
        # PowerShell wraps single results in a dict, not a list
        if isinstance(raw, dict):
            raw = [raw]

        df = pd.DataFrame(raw)
        elapsed = time.perf_counter() - t0
        logger.info("AD query complete: %d users retrieved in %.1fs", len(df), elapsed)
        return df

    except subprocess.TimeoutExpired:
        raise TimeoutError(
            f"AD query exceeded {cfg['AD_QUERY_TIMEOUT_SECONDS']}s timeout. "
            "Consider increasing AD_QUERY_TIMEOUT_SECONDS or checking DC connectivity."
        )
    except json.JSONDecodeError as exc:
        raise ValueError(f"Failed to parse PowerShell JSON output: {exc}")
    finally:
        if tmp_ps1.exists():
            tmp_ps1.unlink()


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 5 ─ EXCEL INPUT LOADING
# ──────────────────────────────────────────────────────────────────────────────

def find_latest_excel(input_folder: str, logger: logging.Logger) -> Path:
    """Auto-detect the most recently modified .xlsx/.xls file in input_folder."""
    folder = Path(input_folder)
    folder.mkdir(parents=True, exist_ok=True)
    files = list(folder.glob("*.xlsx")) + list(folder.glob("*.xls"))
    if not files:
        raise FileNotFoundError(
            f"No Excel files found in '{folder.resolve()}'. "
            "Drop your monthly roster file into that folder and re-run."
        )
    latest = max(files, key=lambda p: p.stat().st_mtime)
    logger.info("Monthly file detected: %s (modified %s)",
                latest.name,
                datetime.fromtimestamp(latest.stat().st_mtime).strftime("%Y-%m-%d %H:%M"))
    return latest


def load_excel_roster(cfg: dict, logger: logging.Logger) -> pd.DataFrame:
    """Load the monthly Excel roster and perform basic normalisation."""
    path = find_latest_excel(cfg["INPUT_FOLDER"], logger)

    df = pd.read_excel(
        path,
        sheet_name=cfg["EXCEL_SHEET_NAME"],
        dtype=str,          # Read everything as string to avoid type coercion
    )

    logger.info("Roster loaded: %d rows, %d columns", len(df), len(df.columns))
    logger.debug("Roster columns: %s", list(df.columns))

    # Normalise column headers: strip whitespace
    df.columns = df.columns.str.strip()

    # Validate required match column exists
    excel_match_col = cfg[cfg["MATCH_EXCEL_FIELD"]]
    if excel_match_col not in df.columns:
        raise ValueError(
            f"Required column '{excel_match_col}' not found in Excel file.\n"
            f"Available columns: {list(df.columns)}\n"
            f"Update EXCEL_COLUMN_USERID in CONFIG to match your file's header."
        )

    return df


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 6 ─ DATA NORMALISATION
# ──────────────────────────────────────────────────────────────────────────────

def normalise_key(series: pd.Series, case_sensitive: bool) -> pd.Series:
    """Strip whitespace, optionally lower-case, replace blank/None with NaN."""
    s = series.astype(str).str.strip()
    if not case_sensitive:
        s = s.str.lower()
    return s.replace({"nan": np.nan, "none": np.nan, "": np.nan})


def normalise_dataframe(df: pd.DataFrame, key_cols: list[str], case_sensitive: bool) -> pd.DataFrame:
    """Apply normalise_key to all listed columns that exist in df."""
    df = df.copy()
    for col in key_cols:
        if col in df.columns:
            df[col] = normalise_key(df[col], case_sensitive)
    return df


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 7 ─ COMPARISON ENGINE
# ──────────────────────────────────────────────────────────────────────────────

def resolve_field(cfg: dict, field_key: str) -> str:
    """Resolve a CONFIG key-of-key to its string value (the actual column name)."""
    return cfg[cfg[field_key]]


def build_comparison(
    ad_df: pd.DataFrame,
    roster_df: pd.DataFrame,
    cfg: dict,
    logger: logging.Logger,
) -> dict[str, pd.DataFrame]:
    """
    Core reconciliation engine.
    Returns a dict of labelled DataFrames ready for Excel output.
    """
    t0 = time.perf_counter()
    cs = cfg["CASE_SENSITIVE_MATCH"]

    # -- Resolve actual column names from CONFIG keys -------------------------
    ad_key_col     = resolve_field(cfg, "MATCH_AD_FIELD")
    excel_key_col  = resolve_field(cfg, "MATCH_EXCEL_FIELD")

    logger.info("Primary match: AD[%s] ↔ Excel[%s]", ad_key_col, excel_key_col)

    # -- Normalise both sides -------------------------------------------------
    ad_norm_col     = "__ad_key__"
    excel_norm_col  = "__excel_key__"

    ad_work = ad_df.copy()
    ad_work[ad_norm_col] = normalise_key(ad_work[ad_key_col], cs)

    roster_work = roster_df.copy()
    roster_work[excel_norm_col] = normalise_key(roster_work[excel_key_col], cs)

    # -- Flag duplicates before merging ---------------------------------------
    ad_dupes     = ad_work[ad_work.duplicated(subset=[ad_norm_col], keep=False) & ad_work[ad_norm_col].notna()]
    roster_dupes = roster_work[roster_work.duplicated(subset=[excel_norm_col], keep=False) & roster_work[excel_norm_col].notna()]

    ad_dupes_out     = ad_dupes.drop(columns=[ad_norm_col])
    roster_dupes_out = roster_dupes.drop(columns=[excel_norm_col])

    duplicates = pd.concat([
        ad_dupes_out.assign(Source="AD"),
        roster_dupes_out.assign(Source="Excel"),
    ], ignore_index=True)

    # -- Flag invalid records (null primary key) -------------------------------
    ad_invalid     = ad_work[ad_work[ad_norm_col].isna()].drop(columns=[ad_norm_col])
    roster_invalid = roster_work[roster_work[excel_norm_col].isna()].drop(columns=[excel_norm_col])

    invalid = pd.concat([
        ad_invalid.assign(Source="AD"),
        roster_invalid.assign(Source="Excel"),
    ], ignore_index=True)

    # -- De-duplicate for comparison (keep first occurrence) ------------------
    ad_deduped     = ad_work.drop_duplicates(subset=[ad_norm_col]).dropna(subset=[ad_norm_col])
    roster_deduped = roster_work.drop_duplicates(subset=[excel_norm_col]).dropna(subset=[excel_norm_col])

    # -- Vectorised set operations on normalised keys -------------------------
    ad_keys     = set(ad_deduped[ad_norm_col])
    excel_keys  = set(roster_deduped[excel_norm_col])

    in_both    = ad_keys & excel_keys
    only_in_ad = ad_keys - excel_keys
    only_in_ex = excel_keys - ad_keys

    # -- Build result frames --------------------------------------------------
    to_remove = (
        ad_deduped[ad_deduped[ad_norm_col].isin(only_in_ad)]
        .drop(columns=[ad_norm_col])
        .reset_index(drop=True)
    )

    to_add = (
        roster_deduped[roster_deduped[excel_norm_col].isin(only_in_ex)]
        .drop(columns=[excel_norm_col])
        .reset_index(drop=True)
    )

    # Matching: merge to get both sides' columns side-by-side
    matching = (
        ad_deduped[ad_deduped[ad_norm_col].isin(in_both)]
        .merge(
            roster_deduped[roster_deduped[excel_norm_col].isin(in_both)],
            left_on=ad_norm_col,
            right_on=excel_norm_col,
            how="inner",
            suffixes=("_AD", "_Excel"),
        )
        .drop(columns=[ad_norm_col, excel_norm_col])
        .reset_index(drop=True)
    )

    elapsed = time.perf_counter() - t0
    logger.info(
        "Comparison complete in %.2fs → Add:%d | Remove:%d | Match:%d | "
        "Dupes:%d | Invalid:%d",
        elapsed, len(to_add), len(to_remove), len(matching),
        len(duplicates), len(invalid),
    )

    # -- Summary metrics ------------------------------------------------------
    summary_data = {
        "Metric": [
            "AD Group Name",
            "Run Timestamp",
            "Total AD Members",
            "Total Roster Entries",
            "─── Matched (in both) ───",
            "Users to ADD (in roster, not in AD)",
            "Users to REMOVE (in AD, not in roster)",
            "─── Data Quality ───",
            "Duplicate Records",
            "Invalid / Blank Key Records",
            "Primary Match Key – AD Field",
            "Primary Match Key – Excel Field",
            "Case Sensitive Matching",
        ],
        "Value": [
            cfg["AD_GROUP_NAME"],
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            len(ad_df),
            len(roster_df),
            "",
            len(to_add),
            len(to_remove),
            "",
            len(duplicates),
            len(invalid),
            ad_key_col,
            excel_key_col,
            str(cs),
        ],
    }

    return {
        "summary":    pd.DataFrame(summary_data),
        "to_add":     to_add,
        "to_remove":  to_remove,
        "matching":   matching,
        "duplicates": duplicates,
        "invalid":    invalid,
        "raw_ad":     ad_df,
    }


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 8 ─ EXCEL REPORT GENERATION
# ──────────────────────────────────────────────────────────────────────────────

# -- Colour palette -----------------------------------------------------------
PALETTE = {
    "header_fill":   "1F4E79",   # Dark navy
    "header_font":   "FFFFFF",   # White
    "add_fill":      "C6EFCE",   # Light green
    "add_font":      "276221",   # Dark green
    "remove_fill":   "FFC7CE",   # Light red
    "remove_font":   "9C0006",   # Dark red
    "match_fill":    "DDEBF7",   # Light blue
    "match_font":    "1F4E79",   # Dark blue
    "dupe_fill":     "FFEB9C",   # Light yellow
    "dupe_font":     "9C6500",   # Dark amber
    "invalid_fill":  "F4CCCC",   # Pale red
    "invalid_font":  "990000",   # Dark red
    "summary_fill":  "E2EFDA",   # Light green summary
    "tab_add":       "70AD47",   # Green tab
    "tab_remove":    "FF0000",   # Red tab
    "tab_match":     "4472C4",   # Blue tab
    "tab_dupe":      "FFC000",   # Amber tab
    "tab_invalid":   "FF7F7F",   # Pink tab
    "tab_raw":       "808080",   # Grey tab
    "tab_summary":   "2E75B6",   # Navy tab
    "alt_row":       "F2F2F2",   # Alternating row fill
}


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _make_font(hex_color: str, bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", color=hex_color, bold=bold, size=size)


def _thin_border() -> Border:
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header_row(ws, num_cols: int, fill_hex: str, font_hex: str):
    """Apply header formatting to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = _make_fill(fill_hex)
        cell.font   = _make_font(font_hex, bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[1].height = 22


def _style_data_rows(ws, num_cols: int, row_fill: str = None, font_hex: str = "000000"):
    """Apply data-row formatting from row 2 onwards."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if col_idx := cell.column:
                if col_idx > num_cols:
                    break
            alt = row[0].row % 2 == 0
            if row_fill:
                cell.fill = _make_fill(row_fill)
            elif alt:
                cell.fill = _make_fill(PALETTE["alt_row"])
            cell.font      = _make_font(font_hex)
            cell.border    = _thin_border()
            cell.alignment = Alignment(vertical="center")


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 50):
    """Set column widths based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def _write_dataframe_to_sheet(ws, df: pd.DataFrame, header_fill: str, header_font: str,
                               row_fill: str = None, font_hex: str = "000000"):
    """Write a DataFrame to a worksheet with full formatting."""
    if df.empty:
        ws.append(["No records in this category."])
        return

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    _style_header_row(ws, len(df.columns), header_fill, header_font)
    _style_data_rows(ws, len(df.columns), row_fill, font_hex)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_fit_columns(ws)


def write_summary_sheet(ws, summary_df: pd.DataFrame, results: dict):
    """Create a visually rich summary sheet."""
    ws.sheet_view.showGridLines = False

    # Title block
    ws["A1"] = "AD GROUP RECONCILIATION REPORT"
    ws["A1"].font      = Font(name="Calibri", size=18, bold=True, color=PALETTE["header_fill"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="808080")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 16

    ws.append([])  # Spacer row

    # Metrics table
    ws.append(["Metric", "Value"])
    header_row = ws.max_row
    ws.cell(header_row, 1).fill = _make_fill(PALETTE["header_fill"])
    ws.cell(header_row, 1).font = _make_font(PALETTE["header_font"], bold=True, size=11)
    ws.cell(header_row, 2).fill = _make_fill(PALETTE["header_fill"])
    ws.cell(header_row, 2).font = _make_font(PALETTE["header_font"], bold=True, size=11)

    colour_map = {
        "Users to ADD":    (PALETTE["add_fill"],    PALETTE["add_font"]),
        "Users to REMOVE": (PALETTE["remove_fill"], PALETTE["remove_font"]),
        "Matched":         (PALETTE["match_fill"],  PALETTE["match_font"]),
        "Duplicate":       (PALETTE["dupe_fill"],   PALETTE["dupe_font"]),
        "Invalid":         (PALETTE["invalid_fill"],PALETTE["invalid_font"]),
    }

    for _, row_data in summary_df.iterrows():
        ws.append([row_data["Metric"], row_data["Value"]])
        r = ws.max_row
        metric_cell = ws.cell(r, 1)
        value_cell  = ws.cell(r, 2)
        metric_cell.border = _thin_border()
        value_cell.border  = _thin_border()
        metric_cell.font   = Font(name="Calibri", size=10)

        if str(row_data["Metric"]).startswith("─"):
            metric_cell.font = Font(name="Calibri", size=10, bold=True, color=PALETTE["header_fill"])
            ws.merge_cells(f"A{r}:B{r}")
            continue

        for keyword, (fill, font) in colour_map.items():
            if keyword in str(row_data["Metric"]):
                for cell in (metric_cell, value_cell):
                    cell.fill = _make_fill(fill)
                    cell.font = _make_font(font, bold=True)
                break

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = "A4"


def generate_excel_report(results: dict, cfg: dict, logger: logging.Logger) -> Path:
    """Assemble the full multi-tab Excel workbook and save it."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    group_safe = re.sub(r"[^\w\-]", "_", cfg["AD_GROUP_NAME"])
    filename   = f"AD_Recon_{group_safe}_{timestamp}.xlsx"
    out_path   = Path(cfg["OUTPUT_FOLDER"]) / filename
    Path(cfg["OUTPUT_FOLDER"]).mkdir(parents=True, exist_ok=True)

    if cfg.get("DRY_RUN"):
        logger.info("[DRY RUN] Would write report to %s", out_path)
        return out_path

    logger.info("Building Excel workbook…")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ── Tab definitions: (sheet_name, df_key, tab_colour, row_fill, font) ───
    tab_config = [
        ("Summary",         "summary",    PALETTE["tab_summary"], None,                      "000000"),
        ("Users_To_Add",    "to_add",     PALETTE["tab_add"],     PALETTE["add_fill"],       PALETTE["add_font"]),
        ("Users_To_Remove", "to_remove",  PALETTE["tab_remove"],  PALETTE["remove_fill"],    PALETTE["remove_font"]),
        ("Matching_Users",  "matching",   PALETTE["tab_match"],   PALETTE["match_fill"],     PALETTE["match_font"]),
        ("Duplicate_Records","duplicates",PALETTE["tab_dupe"],    PALETTE["dupe_fill"],      PALETTE["dupe_font"]),
        ("Invalid_Records", "invalid",    PALETTE["tab_invalid"], PALETTE["invalid_fill"],   PALETTE["invalid_font"]),
        ("Raw_AD_Output",   "raw_ad",     PALETTE["tab_raw"],     None,                      "000000"),
    ]

    for sheet_name, df_key, tab_color, row_fill, font_hex in tab_config:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_color

        if sheet_name == "Summary":
            write_summary_sheet(ws, results["summary"], results)
        else:
            _write_dataframe_to_sheet(
                ws, results[df_key],
                PALETTE["header_fill"], PALETTE["header_font"],
                row_fill, font_hex,
            )

    wb.save(out_path)
    logger.info("Report saved → %s", out_path)
    return out_path


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 9 ─ CSV EXPORT
# ──────────────────────────────────────────────────────────────────────────────

def export_csv_copies(results: dict, out_path: Path, logger: logging.Logger):
    """Write one CSV file per result tab alongside the Excel report."""
    csv_dir = out_path.parent / out_path.stem
    csv_dir.mkdir(parents=True, exist_ok=True)
    for key, df in results.items():
        csv_path = csv_dir / f"{key}.csv"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        logger.debug("CSV written: %s", csv_path)
    logger.info("CSV copies saved → %s/", csv_dir)


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 10 ─ ARCHIVE
# ──────────────────────────────────────────────────────────────────────────────

def archive_old_reports(reports_folder: str, archive_folder: str, logger: logging.Logger):
    """Move any existing .xlsx reports into the archive folder."""
    reports_path = Path(reports_folder)
    archive_path = Path(archive_folder)
    archive_path.mkdir(parents=True, exist_ok=True)

    moved = 0
    for f in reports_path.glob("AD_Recon_*.xlsx"):
        dest = archive_path / f.name
        shutil.move(str(f), dest)
        moved += 1
        logger.debug("Archived: %s → %s", f.name, dest)

    if moved:
        logger.info("Archived %d old report(s) to '%s/'", moved, archive_folder)


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 11 ─ MAIN ORCHESTRATOR
# ──────────────────────────────────────────────────────────────────────────────

def main():
    run_start = time.perf_counter()

    # Logging
    logger = setup_logging(CONFIG["LOG_FOLDER"], CONFIG["LOG_LEVEL"])
    logger.info("=" * 70)
    logger.info("AD Reconciliation Tool — START")
    logger.info("Group    : %s", CONFIG["AD_GROUP_NAME"])
    logger.info("Dry Run  : %s", CONFIG["DRY_RUN"])
    logger.info("=" * 70)

    try:
        # Step 1: Archive previous reports
        if CONFIG.get("ARCHIVE_OLD_REPORTS") and not CONFIG.get("DRY_RUN"):
            archive_old_reports(CONFIG["OUTPUT_FOLDER"], CONFIG["ARCHIVE_FOLDER"], logger)

        # Step 2: Query Active Directory
        logger.info("[1/4] Querying Active Directory…")
        ad_df = query_active_directory(CONFIG, logger)

        # Step 3: Load monthly Excel roster
        logger.info("[2/4] Loading monthly Excel roster…")
        roster_df = load_excel_roster(CONFIG, logger)

        # Step 4: Run comparison
        logger.info("[3/4] Running comparison engine…")
        results = build_comparison(ad_df, roster_df, CONFIG, logger)

        # Step 5: Write output
        logger.info("[4/4] Generating Excel report…")
        out_path = generate_excel_report(results, CONFIG, logger)

        if CONFIG.get("EXPORT_CSV_COPIES") and not CONFIG.get("DRY_RUN"):
            export_csv_copies(results, out_path, logger)

        elapsed = time.perf_counter() - run_start
        logger.info("=" * 70)
        logger.info("COMPLETED SUCCESSFULLY in %.1fs", elapsed)
        logger.info("Report → %s", out_path.resolve() if not CONFIG.get("DRY_RUN") else "DRY RUN")
        logger.info("=" * 70)

    except FileNotFoundError as exc:
        logger.error("FILE NOT FOUND: %s", exc)
        sys.exit(1)
    except TimeoutError as exc:
        logger.error("TIMEOUT: %s", exc)
        sys.exit(2)
    except ValueError as exc:
        logger.error("VALIDATION ERROR: %s", exc)
        sys.exit(3)
    except RuntimeError as exc:
        logger.error("RUNTIME ERROR: %s", exc)
        sys.exit(4)
    except Exception as exc:
        logger.exception("UNEXPECTED ERROR: %s", exc)
        sys.exit(99)


if __name__ == "__main__":
    main()
